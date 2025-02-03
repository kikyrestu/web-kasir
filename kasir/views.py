from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, Min
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from .models import Barang
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.db.models import Q
from django.views.decorators.http import require_POST
from .models import Produk
from .models import Kategori
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json
import logging
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from .models import StoreSettings, ReceiptSettings, SystemSettings
import json
from django.core import serializers
from django.http import FileResponse
import os
from datetime import datetime
from django.db import transaction
from django.db import models
from .models import Transaksi, TransaksiDetail
from django.urls import reverse
import xlsxwriter
from io import BytesIO
from django.db.models.functions import TruncDate
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import datetime
from decimal import Decimal, InvalidOperation
import win32print  # Untuk Windows
# Alternatif untuk Linux: import cups
from PIL import Image
import io
import os
import time
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse
from io import BytesIO
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from datetime import datetime
from decimal import Decimal
import base64
from django.conf import settings
from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce, Cast
from django.db.models.fields import DateField
from django.db import connection
from .models import Module  # Pastikan buat model Module dulu
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from functools import wraps
from django.contrib.auth.models import AnonymousUser
from django.http import Http404
from .models import PPOBSaldo, PPOBSaldoHistory, Module
import decimal
from django.db.models import Sum
from django.db.models.functions import TruncDate, TruncMonth
import traceback
from dateutil.relativedelta import relativedelta
import os
import subprocess
from .models import Transaksi, TransaksiDetail, Produk, OperationalExpense
import pytz

logger = logging.getLogger(__name__)

class BaseView(LoginRequiredMixin, View):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs) if hasattr(super(), 'get_context_data') else {}
        context['active_modules'] = Module.objects.filter(active=True).order_by('order')
        return context

def with_modules(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if isinstance(request.user, AnonymousUser):
            return redirect('login')
            
        # Tambahkan pengecekan hak akses
        if request.user.is_superuser:
            modules = Module.objects.filter(active=True, access_admin=True)
        else:
            modules = Module.objects.filter(active=True, access_kasir=True)
            
        modules = modules.order_by('order')
        
        # Get original response
        response = view_func(request, *args, **kwargs)
        
        # If response is TemplateResponse, update context
        if hasattr(response, 'context_data'):
            response.context_data['active_modules'] = modules
            
        return response
    return wrapper

def get_sales_data(request):
    try:
        days = int(request.GET.get('days', 7))
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days-1)
        
        # Gunakan raw SQL yang terbukti berhasil
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DATE(tanggal) as sale_date, SUM(total) as total_sales
                FROM kasir_transaksi 
                WHERE DATE(tanggal) BETWEEN %s AND %s
                GROUP BY DATE(tanggal)
                ORDER BY sale_date
            """, [start_date, end_date])
            raw_data = cursor.fetchall()

        # Format data untuk chart
        labels = []
        values = []
        
        # Buat dictionary dari hasil query
        sales_dict = {
            date.strftime('%Y-%m-%d'): float(total) 
            for date, total in raw_data
        }
        
        # Isi data untuk setiap hari
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            labels.append(current_date.strftime('%d/%m/%Y'))
            values.append(sales_dict.get(date_str, 0.0))
            current_date += timedelta(days=1)
        
        return JsonResponse({
            'labels': labels,
            'values': values
        })
        
    except Exception as e:
        print(f"Error in get_sales_data: {str(e)}")
        return JsonResponse({
            'error': str(e)
        }, status=500)

@with_modules
def dashboard(request):
    today = timezone.now().date()
    
    # Query untuk transaksi dan pendapatan hari ini
    with connection.cursor() as cursor:
        # Query existing untuk transaksi hari ini
        cursor.execute("""
            SELECT COUNT(*) as total_transactions, COALESCE(SUM(total), 0) as total_income
            FROM kasir_transaksi 
            WHERE DATE(tanggal) = %s
        """, [today])
        
        result = cursor.fetchone()
        today_transactions = result[0]
        today_income = result[1]

        # Query untuk menghitung profit hari ini
        cursor.execute("""
            SELECT COALESCE(SUM(td.qty * (p.h_jual - p.hp_beli)), 0) as total_profit
            FROM kasir_transaksi t
            JOIN kasir_transaksidetail td ON t.id = td.transaksi_id
            JOIN kasir_produk p ON td.produk_id = p.id
            WHERE DATE(t.tanggal) = %s
        """, [today])
        
        today_profit = cursor.fetchone()[0]

        # Query untuk total modal dari stok saat ini
        cursor.execute("""
            SELECT COALESCE(SUM(hp_beli * stok), 0) as total_modal_stok
            FROM kasir_produk
        """)
        total_modal_stok = cursor.fetchone()[0]

        # Query untuk total modal yang sudah berkurang (dari barang terjual)
        cursor.execute("""
            SELECT COALESCE(SUM(p.hp_beli * td.qty), 0) as total_modal_terjual
            FROM kasir_transaksidetail td
            JOIN kasir_produk p ON td.produk_id = p.id
            JOIN kasir_transaksi t ON td.transaksi_id = t.id
        """)
        total_modal_terjual = cursor.fetchone()[0]

    # Data lainnya tetap sama
    BATAS_STOK_MENIPIS = 5
    low_stock_items = Produk.objects.filter(
        stok__lte=BATAS_STOK_MENIPIS
    ).order_by('stok')
    
    total_products = Produk.objects.count()
    low_stock_count = low_stock_items.count()
    
    top_products = Produk.objects.annotate(
        total_terjual=Coalesce(Sum('transaksidetail__qty'), 0)
    ).order_by('-total_terjual')[:5]
    
    recent_transactions = Transaksi.objects.order_by('-tanggal')[:5]
    
    context = {
        'today_transactions': today_transactions,
        'today_income': today_income,
        'today_profit': today_profit,
        'low_stock_items': low_stock_items,
        'total_products': total_products,
        'low_stock_count': low_stock_count,
        'top_products': top_products,
        'recent_transactions': recent_transactions,
        'total_modal_stok': total_modal_stok,
        'total_modal_terjual': total_modal_terjual,
    }
    
    return render(request, 'kasir/dashboard.html', context)

def transaksi(request):
    # Ambil transaksi hari ini dengan range waktu yang tepat
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    transaksi_hari_ini = Transaksi.objects.filter(
        tanggal__range=(today_start, today_end)
    ).prefetch_related(
        'transaksidetail_set',
        'transaksidetail_set__produk'
    ).order_by('-tanggal')  # Urutkan dari yang terbaru
    
    print(f"Debug - Query time range: {today_start} to {today_end}")
    print(f"Debug - Found transactions: {transaksi_hari_ini.count()}")
    
    context = {
        'transaksi_hari_ini': transaksi_hari_ini
    }
    return render(request, 'kasir/transaksi.html', context)

def produk(request):
    # Get all categories first
    categories = Kategori.objects.all().order_by('nama')
    
    search_query = request.GET.get('search', '')
    entries = request.GET.get('entries', '10')
    
    # Pastikan entries adalah angka yang valid
    try:
        entries = int(entries)
    except ValueError:
        entries = 10
    
    products = Produk.objects.annotate(
        total_hp_beli=F('hp_beli') * F('stok'),
        total_h_jual=F('h_jual') * F('stok')
    ).all().order_by('-id')
    
    # Handle search
    if search_query:
        products = products.filter(
            Q(nama_barang__icontains=search_query) |
            Q(kategori__nama__icontains=search_query) |  # Pencarian di nama kategori
            Q(hp_beli__icontains=search_query) |
            Q(h_jual__icontains=search_query) |
            Q(stok__icontains=search_query)
        )
    
    # Handle category filter
    category = request.GET.get('category')
    if category:
        products = products.filter(kategori__nama=category)
    
    # Pagination
    paginator = Paginator(products, entries)
    page = request.GET.get('page', 1)
    
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    context = {
        'products': products,
        'categories': categories,
        'search': search_query,
        'entries': entries,
        'selected_category': category
    }
    
    return render(request, 'kasir/produk.html', context)

def laporan(request):
    """View untuk menampilkan halaman laporan"""
    stok_menipis = Produk.objects.filter(stok__gt=0, stok__lte=10).count()
    stok_habis = Produk.objects.filter(stok=0).count()
    
    # Ambil tanggal hari ini dan 30 hari ke belakang untuk default
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    context = {
        'stok_menipis': stok_menipis,
        'stok_habis': stok_habis,
        'default_start': start_date.strftime('%Y-%m-%d'),
        'default_end': end_date.strftime('%Y-%m-%d')
    }
    return render(request, 'kasir/laporan.html', context)

def export_excel(request):
    try:
        # Ubah Product menjadi Produk
        products = Produk.objects.all().select_related('kategori')
        
        # Siapkan data untuk Excel
        data = []
        for product in products:
            data.append({
                'Nama Barang': product.nama_barang,
                'Kategori': product.kategori.nama if product.kategori else '',
                'Harga Pokok Beli': product.hp_beli,
                'Harga Jual': product.h_jual,
                'Stok': product.stok,
                'Total HP Beli': product.hp_beli * product.stok,  # Hitung total
                'Total H Jual': product.h_jual * product.stok,    # Hitung total
                'Tanggal Terjual': product.tgl_terjual.strftime('%d/%m/%Y') if product.tgl_terjual else '-',
                'Tanggal Stok Masuk': product.tgl_stok_masuk.strftime('%d/%m/%Y') if product.tgl_stok_masuk else '-'
            })
        
        # Sisa kode tetap sama
        df = pd.DataFrame(data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Data Produk')
            
            workbook = writer.book
            worksheet = writer.sheets['Data Produk']
            
            money_format = workbook.add_format({'num_format': '#,##0'})
            
            for idx, col in enumerate(df.columns):
                series = df[col]
                max_len = max(
                    series.astype(str).map(len).max(),
                    len(str(series.name))
                ) + 2
                worksheet.set_column(idx, idx, max_len)
                
                if 'Harga' in col or 'Total' in col:
                    worksheet.set_column(idx, idx, max_len, money_format)
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=data_produk_{timestamp}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Export error: {str(e)}")  # Untuk debugging
        return JsonResponse({
            'status': 'error',
            'message': f'Gagal mengexport data: {str(e)}'
        }, status=500)

# API Endpoints
@csrf_exempt
def product_api(request, product_id):
    try:
        product = Produk.objects.get(id=product_id)
        data = {
            'nama_barang': product.nama_barang,
            'kategori': product.kategori,
            'hp_beli': product.hp_beli,
            'h_jual': product.h_jual,
            'stok': product.stok
        }
        return JsonResponse(data)
    except Produk.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

def import_products(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)
            
            # Debug info
            print("Kolom yang ada di Excel:", df.columns.tolist())
            
            # Pastikan semua kolom yang diperlukan ada
            required_columns = ['nama_barang', 'kategori', 'hp_beli', 'h_jual', 'stok']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
                }, status=400)
            
            success_count = 0
            failed_rows = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Validasi data
                        if pd.isna(row['nama_barang']):
                            raise ValueError("Nama barang tidak boleh kosong")
                        
                        nama_barang = str(row['nama_barang']).strip()
                        
                        # Cek apakah produk sudah ada
                        if Produk.objects.filter(nama_barang=nama_barang).exists():
                            raise ValueError(f"Produk dengan nama '{nama_barang}' sudah ada")
                        
                        # Proses kategori
                        kategori = None
                        if pd.notna(row['kategori']):
                            kategori_nama = str(row['kategori']).strip()
                            if kategori_nama:
                                kategori, _ = Kategori.objects.get_or_create(nama=kategori_nama)
                        
                        # Buat produk baru
                        produk = Produk.objects.create(
                            nama_barang=nama_barang,
                            kategori=kategori,
                            hp_beli=float(row['hp_beli']) if pd.notna(row['hp_beli']) else 0,
                            h_jual=float(row['h_jual']) if pd.notna(row['h_jual']) else 0,
                            stok=int(row['stok']) if pd.notna(row['stok']) else 0,
                            tgl_stok_masuk=timezone.now()  # Tambahkan tanggal stok masuk
                        )
                        
                        success_count += 1
                        
                    except Exception as e:
                        failed_rows.append({
                            'row': index + 2,
                            'error': str(e)
                        })
            
            message = f'Berhasil mengimport {success_count} produk.'
            if failed_rows:
                message += f'\n\nBaris yang gagal:'
                for fail in failed_rows:
                    message += f'\nBaris {fail["row"]}: {fail["error"]}'
            
            return JsonResponse({
                'status': 'success' if success_count > 0 else 'warning',
                'message': message,
                'failed_rows': failed_rows
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=400)

    return JsonResponse({
        'status': 'error',
        'message': 'Method tidak diizinkan atau file tidak ditemukan'
    }, status=400)

def download_template(request):
    """Download template Excel untuk import produk"""
    # Create Excel template
    df = pd.DataFrame(columns=[
        'Nama Barang',
        'Kategori',
        'Harga Pokok Beli',
        'Harga Jual',
        'Stok'
    ])
    
    # Create sample data
    sample_data = {
        'Nama Barang': ['Contoh Produk 1', 'Contoh Produk 2'],
        'Kategori': ['Kategori A', 'Kategori B'],
        'Harga Pokok Beli': [10000, 20000],
        'Harga Jual': [15000, 25000],
        'Stok': [100, 50]
    }
    
    df = pd.DataFrame(sample_data)
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=template_import_produk.xlsx'
    
    # Save to excel
    df.to_excel(response, index=False, engine='openpyxl')
    
    return response

def update_stock(request, product_id):
    if request.method == 'POST':
        try:
            product = Barang.objects.get(no=product_id)
            jumlah = int(request.POST.get('jumlah', 0))
            
            if product.kurangi_stok(jumlah):
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'error': 'Stok tidak mencukupi'}, status=400)
                
        except Barang.DoesNotExist:
            return JsonResponse({'error': 'Produk tidak ditemukan'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@require_POST
def add_product(request):
    try:
        # Validasi input
        nama_barang = request.POST.get('nama_barang')
        if not nama_barang:
            raise ValueError('Nama barang harus diisi')

        # Ambil dan validasi kategori
        kategori_id = request.POST.get('kategori')
        kategori = None
        if kategori_id:
            try:
                kategori = Kategori.objects.get(id=kategori_id)
            except Kategori.DoesNotExist:
                raise ValueError('Kategori tidak valid')

        # Konversi dan validasi harga
        try:
            hp_beli = Decimal(request.POST.get('hp_beli', '0'))
            h_jual = Decimal(request.POST.get('h_jual', '0'))
            stok = int(request.POST.get('stok', '0'))
        except (TypeError, ValueError, InvalidOperation):
            raise ValueError('Format harga atau stok tidak valid')

        if hp_beli < 0 or h_jual < 0 or stok < 0:
            raise ValueError('Harga dan stok tidak boleh negatif')

        # Buat produk baru
        product = Produk.objects.create(
            nama_barang=nama_barang,
            kategori=kategori,
            hp_beli=hp_beli,
            h_jual=h_jual,
            stok=stok
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Produk berhasil ditambahkan',
            'product': {
                'id': product.id,
                'nama_barang': product.nama_barang,
                'kode_barang': product.kode_barang,
                'kategori': product.kategori.nama if product.kategori else None,
                'hp_beli': float(product.hp_beli),
                'h_jual': float(product.h_jual),
                'stok': product.stok
            }
        })
    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Gagal menambahkan produk: {str(e)}'
        }, status=400)

def tambah_produk(request):
    context = {
        'kategoris': Kategori.objects.all()
    }
    return render(request, 'kasir/tambah-produk/tambah-produk.html', context)

@require_POST
def add_kategori(request):
    try:
        nama_kategori = request.POST.get('nama')
        # Cek apakah kategori sudah ada
        kategori, created = Kategori.objects.get_or_create(nama=nama_kategori)
        return JsonResponse({
            'status': 'success',
            'message': 'Kategori berhasil ditambahkan' if created else 'Kategori sudah ada',
            'id': kategori.id,
            'nama': kategori.nama
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

def get_kategoris(request):
    try:
        term = request.GET.get('term', '')
        kategoris = Kategori.objects.filter(nama__icontains=term)
        results = [{'id': k.id, 'text': k.nama} for k in kategoris]
        return JsonResponse({'results': results})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@require_POST
def delete_product(request, product_id):
    if request.method == 'POST':
        try:
            # Gunakan get_object_or_404 untuk handling jika produk tidak ditemukan
            product = get_object_or_404(Produk, id=product_id)
            nama_produk = product.nama_barang
            product.delete()
            messages.success(request, f'Produk {nama_produk} berhasil dihapus')
        except Exception as e:
            messages.error(request, f'Gagal menghapus produk: {str(e)}')
    
    return redirect('kasir:produk')

@require_POST
def delete_products_batch(request):
    try:
        # Ambil ID produk yang akan dihapus dari request
        ids = json.loads(request.POST.get('ids', '[]'))
        
        # Hapus produk berdasarkan ID
        deleted_count = Produk.objects.filter(id__in=ids).delete()[0]
        
        return JsonResponse({
            'status': 'success',
            'message': f'Berhasil menghapus {deleted_count} produk',
            'deleted_count': deleted_count
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@require_POST
def edit_product(request, pk):
    try:
        product = Produk.objects.get(pk=pk)
        
        # Update data produk
        product.nama_barang = request.POST.get('nama_barang')
        kategori_nama = request.POST.get('kategori')
        if kategori_nama:
            kategori, _ = Kategori.objects.get_or_create(nama=kategori_nama)
            product.kategori = kategori
        product.hp_beli = request.POST.get('hp_beli')
        product.h_jual = request.POST.get('h_jual')
        product.stok = request.POST.get('stok')
        
        product.save()
        return JsonResponse({'status': 'success'})
    except Produk.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Produk tidak ditemukan'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def get_printers():
    try:
        # Debug: Print semua printer yang ditemukan
        print("=== Mencari Printer ===")
        printers = []
        for printer in win32print.EnumPrinters(2):  # PRINTER_ENUM_LOCAL | PRINTER_ENUM_CONNECTIONS = 2
            printer_name = printer[2]
            default_printer = win32print.GetDefaultPrinter()
            print(f"Printer ditemukan: {printer_name} (Default: {printer_name == default_printer})")
            
            printers.append({
                'name': printer_name,
                'is_default': printer_name == default_printer
            })
        
        print(f"Total printer ditemukan: {len(printers)}")
        return printers
    except Exception as e:
        print(f"Error saat mencari printer: {str(e)}")
        return []

def settings(request):
    store = StoreSettings.objects.first()
    receipt = ReceiptSettings.objects.first()
    system = SystemSettings.objects.first()
    
    # Get PPOB module status
    ppob_module = Module.objects.filter(path='ppob').first()
    ppob_active = ppob_module.active if ppob_module else False
    
    # Debug: Print hasil get_printers
    available_printers = get_printers()
    print("Printers yang akan ditampilkan:", available_printers)
    
    context = {
        'modules': Module.objects.all(),
        'store': store,
        'receipt': receipt,
        'system': system,   
        'printers': available_printers,
        'ppob_active': ppob_active,
    }
    return render(request, 'kasir/settings.html', context)

@login_required
@require_POST
def update_store_info(request):
    try:
        store_settings = StoreSettings.objects.get(pk=1)
        store_settings.name = request.POST.get('store_name')
        store_settings.address = request.POST.get('address')
        store_settings.phone = request.POST.get('phone')
        store_settings.email = request.POST.get('email')
        
        if 'logo' in request.FILES:
            store_settings.logo = request.FILES['logo']
            
        store_settings.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
@require_POST
def update_receipt_settings(request):
    if request.method == 'POST':
        receipt = ReceiptSettings.objects.first()
        if not receipt:
            receipt = ReceiptSettings()
            
        receipt.header = request.POST.get('receipt_header', '')
        receipt.footer = request.POST.get('receipt_footer', '')
        receipt.paper_size = request.POST.get('paper_size', '58mm')
        receipt.show_logo = request.POST.get('show_logo') == 'on'
        receipt.printer_name = request.POST.get('receipt_printer', '')
        receipt.save()
        
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
@require_POST
def update_system_settings(request):
    try:
        system_settings = SystemSettings.objects.get(pk=1)
        system_settings.currency = request.POST.get('currency')
        system_settings.date_format = request.POST.get('date_format')
        system_settings.timezone = request.POST.get('timezone')
        system_settings.low_stock_threshold = int(request.POST.get('low_stock_threshold', 10))
        system_settings.enable_email_notifications = request.POST.get('enable_email_notifications') == 'true'
        system_settings.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
@require_POST
def backup_data(request):
    try:
        # Create backup directory if it doesn't exist
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
            
        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_{timestamp}.json'
        
        # Serialize data
        data = {
            'store_settings': serializers.serialize('json', StoreSettings.objects.all()),
            'receipt_settings': serializers.serialize('json', ReceiptSettings.objects.all()),
            'system_settings': serializers.serialize('json', SystemSettings.objects.all()),
            'products': serializers.serialize('json', Produk.objects.all()),
            'categories': serializers.serialize('json', Kategori.objects.all()),
            'ppob_saldo': serializers.serialize('json', PPOBSaldo.objects.all()),
            'ppob_history': serializers.serialize('json', PPOBSaldoHistory.objects.all()),
        }
        
        # Create response
        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        print(f"Error in backup_data: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
@require_POST
def restore_data(request):
    try:
        if 'file' not in request.FILES:
            raise ValueError('No file uploaded')
            
        backup_file = request.FILES['file']
        data = json.loads(backup_file.read())
        
        # Restore data for each model
        for model_name, model_data in data.items():
            model_objects = serializers.deserialize('json', model_data)
            for obj in model_objects:
                obj.save()
                
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def kategori_produk(request):
    kategoris = Kategori.objects.annotate(
        produk_count=Count('produk')
    ).order_by('nama')
    
    return render(request, 'kasir/kategori-produk.html', {
        'kategoris': kategoris
    })

@login_required
@require_POST
def add_kategori(request):
    try:
        nama = request.POST.get('nama_kategori')
        deskripsi = request.POST.get('deskripsi')
        
        if not nama:
            return JsonResponse({
                'status': 'error',
                'message': 'Nama kategori harus diisi'
            }, status=400)
            
        Kategori.objects.create(
            nama=nama,
            deskripsi=deskripsi
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Kategori berhasil ditambahkan'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
@require_POST
def update_kategori(request, kategori_id):
    try:
        kategori = Kategori.objects.get(id=kategori_id)
        nama = request.POST.get('nama_kategori')
        deskripsi = request.POST.get('deskripsi')
        
        if not nama:
            return JsonResponse({
                'status': 'error',
                'message': 'Nama kategori harus diisi'
            }, status=400)
            
        kategori.nama = nama
        kategori.deskripsi = deskripsi
        kategori.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Kategori berhasil diperbarui'
        })
    except Kategori.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Kategori tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@require_POST
def hapus_kategori(request):
    try:
        kategori_id = request.POST.get('id')
        kategori = get_object_or_404(Kategori, id=kategori_id)
        nama_kategori = kategori.nama
        
        if kategori.produk_set.exists():
            return JsonResponse({
                'status': 'error',
                'message': f'Kategori {nama_kategori} tidak dapat dihapus karena masih digunakan oleh produk'
            }, status=400)
        
        kategori.delete()
        return JsonResponse({
            'status': 'success',
            'message': f'Kategori {nama_kategori} berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@require_POST
def import_kategori_from_produk(request):
    """Import kategori dari file Excel produk"""
    if request.FILES.get('file'):
        try:
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)
            
            # Debug info
            print(f"Membaca file: {excel_file.name}")
            print(f"Kolom yang tersedia: {df.columns.tolist()}")
            
            # Ambil unique kategori dari kolom kategori
            if 'kategori' not in df.columns:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Kolom kategori tidak ditemukan dalam file'
                }, status=400)
            
            kategoris = df['kategori'].dropna().unique()
            print(f"Kategori yang ditemukan: {kategoris}")
            
            success_count = 0
            with transaction.atomic():
                for kategori_nama in kategoris:
                    # Bersihkan nama kategori
                    kategori_nama = str(kategori_nama).strip()
                    if kategori_nama:
                        # Buat kategori jika belum ada
                        kategori, created = Kategori.objects.get_or_create(
                            nama=kategori_nama,
                            defaults={'deskripsi': ''}
                        )
                        if created:
                            success_count += 1
                            print(f"Berhasil membuat kategori: {kategori_nama}")
            
            return JsonResponse({
                'status': 'success',
                'message': f'Berhasil mengimport {success_count} kategori baru'
            })
            
        except Exception as e:
            print(f"Error saat import kategori: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': f'Terjadi kesalahan: {str(e)}'
            }, status=400)
    
    return JsonResponse({
        'status': 'error',
        'message': 'File tidak ditemukan'
    }, status=400)

@require_http_methods(["GET"])
def get_product(request, pk):
    try:
        product = Produk.objects.get(pk=pk)
        data = {
            'nama_barang': product.nama_barang,
            'kategori': product.kategori.nama if product.kategori else '',
            'hp_beli': product.hp_beli,
            'h_jual': product.h_jual,
            'stok': product.stok
        }
        return JsonResponse(data)
    except Produk.DoesNotExist:
        return JsonResponse({'error': 'Produk tidak ditemukan'}, status=404)

@require_POST
def tambah_kategori(request):
    try:
        nama_kategori = request.POST.get('nama_kategori')
        
        if not nama_kategori:
            return JsonResponse({
                'status': 'error',
                'message': 'Nama kategori tidak boleh kosong'
            }, status=400)
            
        kategori = Kategori.objects.create(nama=nama_kategori)
        return JsonResponse({
            'status': 'success',
            'message': 'Kategori berhasil ditambahkan'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@require_POST
def edit_kategori(request, pk):
    try:
        kategori = Kategori.objects.get(pk=pk)
        kategori.nama = request.POST.get('nama')
        kategori.save()
        return JsonResponse({'status': 'success', 'message': 'Kategori berhasil diupdate'})
    except Kategori.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Kategori tidak ditemukan'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
def delete_kategori(request, pk):
    try:
        kategori = Kategori.objects.get(pk=pk)
        kategori.delete()
        return JsonResponse({'status': 'success', 'message': 'Kategori berhasil dihapus'})
    except Kategori.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Kategori tidak ditemukan'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def search_products(request):
    search_query = request.GET.get('search', '')
    # Pastikan entries diambil dari request, default 10
    entries = request.GET.get('entries', '10')
    page = request.GET.get('page', '1')
    
    print(f"Debug - Entries: {entries}, Page: {page}, Search: {search_query}")  # Debug log
    
    products = Produk.objects.all().order_by('nama_barang')
    
    if search_query:
        products = products.filter(
            Q(nama_barang__icontains=search_query) |
            Q(kategori__nama__icontains=search_query)
        )
    
    # Hitung total
    for product in products:
        product.total_hp_beli = product.hp_beli * product.stok
        product.total_h_jual = product.h_jual * product.stok
    
    # Pagination dengan entries yang benar
    try:
        entries = int(entries)
    except ValueError:
        entries = 10
        
    paginator = Paginator(products, entries)
    
    try:
        products = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        products = paginator.page(1)
    
    context = {
        'products': products,
        'search': search_query,
        'entries': str(entries),  # Konversi ke string untuk template
        'selected_entries': entries,
        'current_page': page
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string(
            'kasir/includes/product_table.html',
            context,
            request=request
        )
        
        pagination = render_to_string(
            'kasir/includes/pagination.html',
            context,
            request=request
        )
        
        return JsonResponse({
            'html': html,
            'pagination': pagination
        })
    
    return render(request, 'kasir/produk.html', context)

@require_POST
@transaction.atomic
def process_transaction(request):
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        total = data.get('total')
        bayar = data.get('bayar')
        kembalian = data.get('kembalian')

        print("=== DEBUG TRANSAKSI ===")
        print(f"Items: {items}")
        print(f"Total: {total}")
        print(f"Bayar: {bayar}")
        print(f"Kembalian: {kembalian}")

        if not items:
            return JsonResponse({
                'status': 'error',
                'message': 'Keranjang kosong'
            }, status=400)

        # Buat transaksi baru
        transaksi = Transaksi.objects.create(
            total=total,
            bayar=bayar,
            kembalian=kembalian,
            tanggal=timezone.now()
        )

        # Proses setiap item
        for item in items:
            try:
                produk = Produk.objects.select_for_update().get(id=item['id'])
                print(f"Processing product: {produk.nama_barang} (ID: {produk.id})")

                if produk.stok < item['qty']:
                    raise ValueError(f"Stok {produk.nama_barang} tidak mencukupi")

                # Buat detail transaksi dengan produk_id yang benar
                TransaksiDetail.objects.create(
                    transaksi=transaksi,
                    produk_id=produk.id,  # Gunakan produk.id langsung
                    qty=item['qty'],
                    harga=item['price']
                )

                # Update stok
                produk.stok -= item['qty']
                produk.save()
                print(f"Updated stock for {produk.nama_barang} (ID: {produk.id}): {produk.stok}")

            except Produk.DoesNotExist:
                raise ValueError(f"Produk dengan ID {item['id']} tidak ditemukan")

        print("Transaction completed successfully")
        return JsonResponse({
            'status': 'success',
            'message': 'Transaksi berhasil',
            'transaksi_id': transaksi.id
        })

    except ValueError as e:
        print(f"ValueError: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

def print_receipt(request, transaksi_id):
    try:
        transaksi = Transaksi.objects.get(id=transaksi_id)
        details = TransaksiDetail.objects.filter(transaksi=transaksi).select_related('produk')
        store_settings = StoreSettings.objects.first()
        receipt_settings = ReceiptSettings.objects.first()

        printer_name = receipt_settings.printer_name if receipt_settings else None
        if not printer_name:
            raise Exception("Printer belum dipilih di pengaturan")

        # Inisialisasi printer
        ESC = b'\x1B'
        GS = b'\x1D'
        INIT = ESC + b'@'          # Initialize printer
        ALIGN_CENTER = ESC + b'a1'  # Center alignment
        ALIGN_LEFT = ESC + b'a0'    # Left alignment
        BOLD_ON = ESC + b'E1'      # Bold text on
        BOLD_OFF = ESC + b'E0'     # Bold text off
        DOUBLE_ON = GS + b'!1'     # Double size text
        DOUBLE_OFF = GS + b'!0'    # Normal size text
        FEED_AND_CUT = GS + b'V\x42\x00'  # Feed and cut paper
        LINE_FEED = b'\x0A'        # Line feed

        # Buat string untuk struk
        receipt_parts = []
        
        # Header
        receipt_parts.extend([
            INIT,
            ALIGN_CENTER,
            DOUBLE_ON,  # Nama toko ukuran double
            store_settings.name.center(16).encode('ascii', 'replace') + LINE_FEED,
            DOUBLE_OFF,
            store_settings.address.center(32).encode('ascii', 'replace') + LINE_FEED if store_settings else b'',
            (f"Telp: {store_settings.phone}").center(32).encode('ascii', 'replace') + LINE_FEED if store_settings and store_settings.phone else b'',
            b'=' * 32 + LINE_FEED,  # Garis pembatas tebal
            ALIGN_LEFT,
        ])
        
        # Info transaksi
        receipt_parts.extend([
            BOLD_ON,
            b'No   : ' + f"#{transaksi.id}".encode('ascii', 'replace') + LINE_FEED,
            b'Tgl  : ' + f"{transaksi.tanggal.strftime('%d/%m/%y %H:%M')}".encode('ascii', 'replace') + LINE_FEED,
            b'Kasir: ' + request.user.username.encode('ascii', 'replace') + LINE_FEED,
            BOLD_OFF,
            b'-' * 32 + LINE_FEED,
        ])
        
        # Detail items
        for item in details:
            nama = item.produk.nama_barang if item.produk else 'Unknown'
            if len(nama) > 20:
                nama = nama[:17] + '...'
            
            qty_str = str(item.qty).rjust(3)
            harga_str = f"{item.harga:,.0f}".rjust(10)
            subtotal = item.qty * item.harga
            subtotal_str = f"{subtotal:,.0f}".rjust(10)
            
            receipt_parts.extend([
                nama.encode('ascii', 'replace') + LINE_FEED,
                f"{qty_str} x {harga_str} = {subtotal_str}".encode('ascii', 'replace') + LINE_FEED,
            ])
        
        # Footer
        receipt_parts.extend([
            b'-' * 32 + LINE_FEED,
            BOLD_ON,
            b'TOTAL    : Rp ' + f"{transaksi.total:,.0f}".rjust(10).encode('ascii', 'replace') + LINE_FEED,
            b'BAYAR    : Rp ' + f"{transaksi.bayar:,.0f}".rjust(10).encode('ascii', 'replace') + LINE_FEED,
            b'KEMBALI  : Rp ' + f"{transaksi.kembalian:,.0f}".rjust(10).encode('ascii', 'replace') + LINE_FEED,
            BOLD_OFF,
            b'=' * 32 + LINE_FEED,  # Garis pembatas tebal
            ALIGN_CENTER,
            b"Terima Kasih Atas Kunjungan Anda" + LINE_FEED,
            b"Barang Yang Sudah Dibeli" + LINE_FEED,
            b"Tidak Dapat Ditukar/Dikembalikan" + LINE_FEED,
            LINE_FEED + LINE_FEED + LINE_FEED,  # 3 baris kosong sebelum potong
            FEED_AND_CUT  # Cut paper
        ])

        # Gabungkan semua bagian struk
        receipt_bytes = b''.join(receipt_parts)

        try:
            # Print using default Windows printer settings
            handle = win32print.OpenPrinter(printer_name)
            try:
                win32print.StartDocPrinter(handle, 1, ("Struk", None, "RAW"))
                win32print.StartPagePrinter(handle)
                win32print.WritePrinter(handle, receipt_bytes)
                win32print.EndPagePrinter(handle)
                win32print.EndDocPrinter(handle)
            finally:
                win32print.ClosePrinter(handle)
            
            return JsonResponse({'status': 'success', 'message': 'Struk berhasil dicetak'})
            
        except Exception as printer_error:
            error_msg = f"Gagal mencetak: {str(printer_error)}"
            if "Access is denied" in str(printer_error):
                error_msg = "Akses ke printer ditolak. Pastikan printer sudah dinyalakan dan terhubung."
            elif "Invalid printer name" in str(printer_error):
                error_msg = "Nama printer tidak valid. Periksa pengaturan printer."
            
            return JsonResponse({'status': 'error', 'message': error_msg}, status=400)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def test_print(request):
    try:
        printer_name = request.POST.get('printer_name')
        if not printer_name:
            return JsonResponse({'status': 'error', 'message': 'Nama printer tidak ditemukan'})

        store_settings = StoreSettings.objects.first()

        # Inisialisasi printer
        ESC = b'\x1B'
        GS = b'\x1D'
        INIT = ESC + b'@'          # Initialize printer
        ALIGN_CENTER = ESC + b'a1'  # Center alignment
        ALIGN_LEFT = ESC + b'a0'    # Left alignment
        BOLD_ON = ESC + b'E1'      # Bold text on
        BOLD_OFF = ESC + b'E0'     # Bold text off
        DOUBLE_ON = GS + b'!1'     # Double size text
        DOUBLE_OFF = GS + b'!0'    # Normal size text
        FEED_AND_CUT = GS + b'V\x42\x00'  # Feed and cut paper
        LINE_FEED = b'\x0A'        # Line feed

        # Buat string untuk test print
        test_parts = []
        
        test_parts.extend([
            INIT,
            ALIGN_CENTER,
            DOUBLE_ON,
            store_settings.name.center(16).encode('ascii', 'replace') + LINE_FEED,
            DOUBLE_OFF,
            store_settings.address.center(32).encode('ascii', 'replace') + LINE_FEED if store_settings else b'',
            (f"Telp: {store_settings.phone}").center(32).encode('ascii', 'replace') + LINE_FEED if store_settings and store_settings.phone else b'',
            b'=' * 32 + LINE_FEED,
            BOLD_ON,
            b'TEST PRINT' + LINE_FEED,
            BOLD_OFF,
            b'-' * 32 + LINE_FEED,
            ALIGN_LEFT,
            b'Printer : ' + printer_name.encode('ascii', 'replace') + LINE_FEED,
            b'Waktu  : ' + timezone.now().strftime('%d/%m/%y %H:%M').encode('ascii', 'replace') + LINE_FEED,
            b'-' * 32 + LINE_FEED,
            ALIGN_CENTER,
            BOLD_ON,
            b'PRINTER SIAP DIGUNAKAN' + LINE_FEED,
            BOLD_OFF,
            b'=' * 32 + LINE_FEED,
            LINE_FEED + LINE_FEED + LINE_FEED,  # 3 baris kosong sebelum potong
            FEED_AND_CUT  # Cut paper
        ])

        # Gabungkan semua bagian
        test_bytes = b''.join(test_parts)

        try:
            # Print using default Windows printer settings
            handle = win32print.OpenPrinter(printer_name)
            try:
                win32print.StartDocPrinter(handle, 1, ("Test Print", None, "RAW"))
                win32print.StartPagePrinter(handle)
                win32print.WritePrinter(handle, test_bytes)
                win32print.EndPagePrinter(handle)
                win32print.EndDocPrinter(handle)
            finally:
                win32print.ClosePrinter(handle)

            return JsonResponse({'status': 'success', 'message': 'Test print berhasil'})

        except Exception as printer_error:
            error_msg = f"Gagal test print: {str(printer_error)}"
            if "Access is denied" in str(printer_error):
                error_msg = "Akses ke printer ditolak. Pastikan printer sudah dinyalakan dan terhubung."
            elif "Invalid printer name" in str(printer_error):
                error_msg = "Nama printer tidak valid. Periksa pengaturan printer."
            
            return JsonResponse({'status': 'error', 'message': error_msg}, status=400)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error saat test print: {str(e)}'}, status=400)

def get_sales_report(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Query transaksi
        query = Transaksi.objects.filter(
            tanggal__gte=f"{start_date} 00:00:00",
            tanggal__lte=f"{end_date} 23:59:59"
        )
        
        # Hitung total penjualan
        total_sales = sum(t.total for t in query.all())
        
        # Hitung total profit dari selisih harga jual dan hpp
        total_profit = TransaksiDetail.objects.filter(
            transaksi__in=query
        ).aggregate(
            total=Sum(F('qty') * (F('barang__h_jual') - F('barang__hp_beli')))
        )['total'] or Decimal('0')
        
        # Hitung beban operasional
        fixed_salary = Decimal('2000000')  # Beban gaji tetap 2jt
        
        total_expense = OperationalExpense.objects.filter(
            date__range=[start_date, end_date]
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0') + fixed_salary
        
        # Hitung profit bersih
        net_profit = total_profit - total_expense
        
        return JsonResponse({
            'total_sales': float(total_sales),
            'total_profit': float(total_profit),
            'net_profit': float(net_profit)
        })
        
    except Exception as e:
        print(f"Error dalam get_sales_report: {str(e)}")
        return JsonResponse({
            'total_sales': 0,
            'total_profit': 0,
            'net_profit': 0,
            'error': str(e)
        })

def get_profit_loss_report(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Convert string dates to datetime objects
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Hitung Penjualan
        sales = Transaksi.objects.filter(
            tanggal__range=[start_date, end_date]
        )
        
        gross_sales = sales.aggregate(
            total=Sum('total')
        )['total'] or Decimal('0')
        
        # Hitung Retur Penjualan (jika ada sistem retur)
        sales_returns = Decimal('0')
        
        # Hitung Penjualan Bersih
        net_sales = gross_sales - sales_returns
        
        # Hitung HPP
        sale_items = TransaksiDetail.objects.filter(
            transaksi__in=sales
        ).aggregate(
            total_hpp=Sum(F('qty') * F('barang__hp_beli'))
        )
        total_cogs = sale_items['total_hpp'] or Decimal('0')
        
        # Hitung Laba Kotor
        gross_profit = net_sales - total_cogs
        
        # Hitung Beban Operasional
        # Tambahkan beban gaji tetap 2jt
        fixed_salary = Decimal('2000000')  # Beban gaji tetap 2jt
        
        expenses = OperationalExpense.objects.filter(
            date__range=[start_date, end_date]
        )
        
        # Gunakan beban gaji tetap
        salary_expense = fixed_salary
        
        utility_expense = expenses.filter(
            expense_type='UTILITY'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        other_expense = expenses.filter(
            expense_type='OTHER'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        total_expense = salary_expense + utility_expense + other_expense
        
        # Hitung Laba Bersih
        net_profit = gross_profit - total_expense
        
        # Inventory values
        beginning_inventory = Produk.objects.aggregate(
            total=Sum(F('stok') * F('hp_beli'))
        )['total'] or Decimal('0')
        
        ending_inventory = beginning_inventory  # Perlu disesuaikan dengan logika bisnis
        
        # Purchases (implementasikan sesuai kebutuhan)
        purchases = Decimal('0')
        
        return JsonResponse({
            'gross_sales': float(gross_sales),
            'sales_returns': float(sales_returns),
            'net_sales': float(net_sales),
            'beginning_inventory': float(beginning_inventory),
            'purchases': float(purchases),
            'ending_inventory': float(ending_inventory),
            'total_cogs': float(total_cogs),
            'gross_profit': float(gross_profit),
            'salary_expense': float(salary_expense),  # Akan menampilkan 2jt
            'utility_expense': float(utility_expense),
            'other_expense': float(other_expense),
            'total_expense': float(total_expense),
            'net_profit': float(net_profit)
        })
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=400)

def export_profit_loss(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    format = request.GET.get('format', 'xlsx')
    
    if format == 'xlsx':
        return export_profit_loss_excel(request, start_date, end_date)
    else:
        return export_profit_loss_pdf(request, start_date, end_date)

def export_profit_loss_excel(request, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Laba Rugi"
    
    # Styling
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True)
    header_style.fill = PatternFill("solid", fgColor="CCCCCC")
    
    # Title
    ws['A1'] = 'LAPORAN LABA RUGI'
    ws['A2'] = f'Periode: {start_date} s/d {end_date}'
    
    # Get data
    data = get_profit_loss_data(start_date, end_date)
    
    # Write data
    current_row = 4
    
    # Pendapatan
    ws.cell(row=current_row, column=1, value="PENDAPATAN").style = header_style
    current_row += 1
    ws.cell(row=current_row, column=1, value="Penjualan Kotor")
    ws.cell(row=current_row, column=2, value=data['gross_sales'])
    
    # ... tambahkan semua komponen laba rugi ...
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Laporan_Laba_Rugi_{start_date}_{end_date}.xlsx'
    wb.save(response)
    return response

def export_profit_loss_pdf(request, start_date, end_date):
    # Get data
    data = get_profit_loss_data(start_date, end_date)
    
    # Render template
    template = get_template('kasir/profit_loss_pdf.html')
    html = template.render({
        'data': data,
        'start_date': start_date,
        'end_date': end_date
    })
    
    # Create PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Laporan_Laba_Rugi_{start_date}_{end_date}.pdf'
        return response
    
    return HttpResponse('Error Generating PDF', status=500)

def export_stock_report(request):
    # ... kode export_stock_report yang sudah ada ...
    pass

def export_sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    try:
        # Create workbook and sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Laporan Detail"
        ws2 = wb.create_sheet("Laporan Ringkas")
        ws3 = wb.create_sheet("Ringkasan Modal")
        
        # Headers untuk sheet 3 (Ringkasan Modal)
        headers3 = [
            'Keterangan',
            'Nominal'
        ]
        ws3.append(headers3)
        
        # Get total modal awal (semua produk)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COALESCE(SUM(hp_beli * stok), 0) as total_modal_awal
                FROM kasir_produk
            """)
            total_modal_awal = Decimal(str(cursor.fetchone()[0]))
            
            # Get total modal yang terjual pada periode ini
            cursor.execute("""
                SELECT COALESCE(SUM(td.qty * p.hp_beli), 0) as total_modal_terjual
                FROM kasir_transaksidetail td
                JOIN kasir_produk p ON td.produk_id = p.id
                JOIN kasir_transaksi t ON td.transaksi_id = t.id
                WHERE DATE(t.tanggal) BETWEEN %s AND %s
            """, [start_date, end_date])
            total_modal_terjual = Decimal(str(cursor.fetchone()[0]))
            
            # Hitung sisa modal
            sisa_modal = total_modal_awal - total_modal_terjual
            
            # Tambahkan data ke sheet 3
            ws3.append(['Total Modal Awal (Semua Produk)', float(total_modal_awal)])
            ws3.append(['Total Modal Terjual (Periode Ini)', float(total_modal_terjual)])
            ws3.append(['Sisa Modal', float(sisa_modal)])
            
            # Style sheet 3
            for row in ws3.iter_rows(min_row=1, max_row=4, min_col=1, max_col=2):
                for cell in row:
                    if row[0].row == 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                    else:
                        if cell.column == 2:  # Kolom nominal
                            cell.number_format = '#,##0'
                        if row[0].row == 4:  # Total row
                            cell.font = Font(bold=True)
            
            # Set column width untuk sheet 3
            ws3.column_dimensions['A'].width = 40
            ws3.column_dimensions['B'].width = 20
        
        # Headers untuk sheet 1 (Detail)
        headers1 = [
            'No',
            'Nama Produk',
            'Tanggal Transaksi',
            'Jumlah Terjual',
            'Harga Satuan',
            'HPP Satuan',
            'Total Penjualan',
            'Total HPP',
            'Laba',
            'Stok Awal',
            'Sisa Stok',
            'Total HPP Awal',
            'Sisa HPP'
        ]
        
        # Headers untuk sheet 2 (Ringkas)
        headers2 = [
            'No',
            'Nama Produk',
            'Total Terjual',
            'Harga Satuan',
            'HPP Satuan',
            'Total Penjualan',
            'Total HPP',
            'Total Laba',
            'Stok Awal',
            'Sisa Stok',
            'Total HPP Awal',
            'Sisa HPP'
        ]
        
        ws1.append(headers1)
        ws2.append(headers2)
        
        # Get data untuk sheet 1 (Detail)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    p.nama_barang,
                    t.tanggal,
                    td.qty,
                    td.harga,
                    p.hp_beli,
                    p.stok,
                    (
                        SELECT COALESCE(SUM(td2.qty), 0)
                        FROM kasir_transaksidetail td2
                        JOIN kasir_transaksi t2 ON td2.transaksi_id = t2.id
                        WHERE td2.produk_id = p.id
                        AND t2.tanggal < %s
                    ) as total_terjual_sebelumnya
                FROM kasir_transaksidetail td
                JOIN kasir_transaksi t ON td.transaksi_id = t.id
                LEFT JOIN kasir_produk p ON td.produk_id = p.id
                WHERE DATE(t.tanggal) BETWEEN %s AND %s
                ORDER BY t.tanggal DESC, td.id ASC
            """, [start_date, start_date, end_date])
            
            sales_data = cursor.fetchall()
            
            # Get data untuk sheet 2 (Ringkas)
            cursor.execute("""
                SELECT 
                    p.nama_barang,
                    SUM(td.qty) as total_qty,
                    td.harga,
                    p.hp_beli,
                    p.stok,
                    (
                        SELECT COALESCE(SUM(td2.qty), 0)
                        FROM kasir_transaksidetail td2
                        JOIN kasir_transaksi t2 ON td2.transaksi_id = t2.id
                        WHERE td2.produk_id = p.id
                        AND t2.tanggal < %s
                    ) as total_terjual_sebelumnya,
                    p.id
                FROM kasir_transaksidetail td
                JOIN kasir_transaksi t ON td.transaksi_id = t.id
                LEFT JOIN kasir_produk p ON td.produk_id = p.id
                WHERE DATE(t.tanggal) BETWEEN %s AND %s
                GROUP BY p.id, p.nama_barang, td.harga, p.hp_beli, p.stok
                ORDER BY p.nama_barang
            """, [start_date, start_date, end_date])
            
            summary_data = cursor.fetchall()
        
        # Proses data untuk sheet 1 (Detail)
        total_penjualan = Decimal('0')
        total_hpp = Decimal('0')
        total_laba = Decimal('0')
        total_hpp_awal = Decimal('0')
        total_sisa_hpp = Decimal('0')
        
        for idx, item in enumerate(sales_data, 1):
            nama_produk = item[0] if item[0] else f"Produk dengan harga {item[3]:,.0f}"
            tanggal = item[1]
            qty = item[2]
            harga_satuan = Decimal(str(item[3]))
            hpp_satuan = Decimal(str(item[4])) if item[4] else Decimal('0')
            sisa_stok = item[5] if item[5] is not None else 0
            total_terjual_sebelumnya = item[6]
            
            stok_awal = sisa_stok + qty + total_terjual_sebelumnya
            total_jual = qty * harga_satuan
            total_hpp_produk = qty * hpp_satuan
            laba = total_jual - total_hpp_produk
            total_hpp_awal_produk = stok_awal * hpp_satuan
            sisa_hpp_produk = sisa_stok * hpp_satuan
            
            total_penjualan += total_jual
            total_hpp += total_hpp_produk
            total_laba += laba
            total_hpp_awal += total_hpp_awal_produk
            total_sisa_hpp += sisa_hpp_produk
            
            if isinstance(tanggal, str):
                tanggal = datetime.strptime(tanggal, '%Y-%m-%d %H:%M:%S.%f')
            
            row1 = [
                idx,
                nama_produk,
                tanggal.strftime('%d/%m/%Y %H:%M'),
                qty,
                float(harga_satuan),
                float(hpp_satuan),
                float(total_jual),
                float(total_hpp_produk),
                float(laba),
                stok_awal,
                sisa_stok,
                float(total_hpp_awal_produk),
                float(sisa_hpp_produk)
            ]
            ws1.append(row1)
        
        # Proses data untuk sheet 2 (Ringkas)
        total_ringkas = {
            'penjualan': Decimal('0'),
            'hpp': Decimal('0'),
            'laba': Decimal('0'),
            'hpp_awal': Decimal('0'),
            'sisa_hpp': Decimal('0')
        }
        
        for idx, item in enumerate(summary_data, 1):
            nama_produk = item[0] if item[0] else f"Produk dengan harga {item[2]:,.0f}"
            total_qty = item[1]
            harga_satuan = Decimal(str(item[2]))
            hpp_satuan = Decimal(str(item[3])) if item[3] else Decimal('0')
            sisa_stok = item[4] if item[4] is not None else 0
            total_terjual_sebelumnya = item[5]
            
            stok_awal = sisa_stok + total_qty + total_terjual_sebelumnya
            total_jual = total_qty * harga_satuan
            total_hpp_produk = total_qty * hpp_satuan
            total_laba = total_jual - total_hpp_produk
            total_hpp_awal_produk = stok_awal * hpp_satuan
            sisa_hpp_produk = sisa_stok * hpp_satuan
            
            total_ringkas['penjualan'] += total_jual
            total_ringkas['hpp'] += total_hpp_produk
            total_ringkas['laba'] += total_laba
            total_ringkas['hpp_awal'] += total_hpp_awal_produk
            total_ringkas['sisa_hpp'] += sisa_hpp_produk
            
            row2 = [
                idx,
                nama_produk,
                total_qty,
                float(harga_satuan),
                float(hpp_satuan),
                float(total_jual),
                float(total_hpp_produk),
                float(total_laba),
                stok_awal,
                sisa_stok,
                float(total_hpp_awal_produk),
                float(sisa_hpp_produk)
            ]
            ws2.append(row2)
        
        # Add total rows
        total_row1 = len(sales_data) + 2
        total_row2 = len(summary_data) + 2
        
        # Total untuk sheet 1
        ws1.append([
            '',
            'TOTAL',
            '',
            '',
            '',
            '',
            f'=SUM(G2:G{total_row1-1})',
            f'=SUM(H2:H{total_row1-1})',
            f'=SUM(I2:I{total_row1-1})',
            '',
            '',
            f'=SUM(L2:L{total_row1-1})',
            f'=SUM(M2:M{total_row1-1})'
        ])
        
        # Total untuk sheet 2
        ws2.append([
            '',
            'TOTAL',
            f'=SUM(C2:C{total_row2-1})',
            '',
            '',
            f'=SUM(F2:F{total_row2-1})',
            f'=SUM(G2:G{total_row2-1})',
            f'=SUM(H2:H{total_row2-1})',
            '',
            '',
            f'=SUM(K2:K{total_row2-1})',
            f'=SUM(L2:L{total_row2-1})'
        ])
        
        # Style kedua sheet
        for ws in [ws1, ws2]:
            # Set column widths
            column_widths = [8, 40, 20, 10, 15, 15, 15, 15, 15, 12, 12, 15, 15] if ws == ws1 else [8, 40, 15, 15, 15, 15, 15, 15, 12, 12, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Format currency columns
            currency_columns = ['E', 'F', 'G', 'H', 'I', 'L', 'M'] if ws == ws1 else ['D', 'E', 'F', 'G', 'H', 'K', 'L']
            max_row = total_row1 if ws == ws1 else total_row2
            for col in currency_columns:
                for row in range(2, max_row + 1):
                    cell = ws[f'{col}{row}']
                    cell.number_format = '#,##0'
            
            # Style total row
            total_row = total_row1 if ws == ws1 else total_row2
            total_cells = ws[total_row]
            for cell in total_cells:
                cell.font = Font(bold=True)
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Laporan_Penjualan_{start_date}_sd_{end_date}.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error in export_sales_report: {str(e)}")
        traceback.print_exc()
        return HttpResponse(f"Error: {str(e)}", status=500)

def search_products_transaksi(request):
    search = request.GET.get('search', '')
    if search:
        products = Produk.objects.select_related('kategori').filter(
            Q(nama_barang__icontains=search)  # Hapus filter barcode karena field tidak ada
        ).values(
            'id', 
            'nama_barang', 
            'h_jual', 
            'stok',
            'kategori__nama'  # Ambil nama kategori
        )[:10]

        # Format data untuk response
        products_list = [{
            'id': p['id'],
            'nama_barang': p['nama_barang'],
            'h_jual': float(p['h_jual']),
            'stok': p['stok'],
            'kategori': p['kategori__nama'] or '-'  # Gunakan '-' jika kategori None
        } for p in products]

        return JsonResponse({
            'status': 'success',
            'products': products_list
        })
    
    return JsonResponse({
        'status': 'error',
        'message': 'Query pencarian kosong'
    })

def process_logo_for_printer(logo_path, max_width=200):  # Ubah ke 200 dots (sekitar 25mm)
    try:
        # Buka gambar
        img = Image.open(logo_path)
        
        # Convert ke grayscale
        img = img.convert('L')
        
        # Resize dengan mempertahankan aspect ratio
        # Batasi tinggi maksimal juga
        max_height = 100  # Sekitar 12mm
        
        # Hitung rasio resize
        width_ratio = max_width / float(img.size[0])
        height_ratio = max_height / float(img.size[1])
        ratio = min(width_ratio, height_ratio)
        
        # Hitung dimensi baru
        new_width = int(float(img.size[0]) * ratio)
        new_height = int(float(img.size[1]) * ratio)
        
        # Resize gambar
        img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
        
        # Convert ke hitam putih murni (dithering)
        img = img.convert('1')
        
        # Convert ke bytes untuk ESC/POS
        buffer = io.BytesIO()
        img.save(buffer, format='BMP')
        
        # Generate ESC/POS commands untuk gambar
        width_low = new_width & 0xff
        width_high = (new_width >> 8) & 0xff
        height_low = new_height & 0xff
        height_high = (new_height >> 8) & 0xff
        
        commands = []
        commands.append(b'\x1D\x76\x30\x00')  # GS v 0 \0
        commands.append(bytes([width_low, width_high]))  # width
        commands.append(bytes([height_low, height_high]))  # height
        commands.append(buffer.getvalue()[62:])  # skip BMP header
        
        return b''.join(commands)
        
    except Exception as e:
        print(f"Error processing logo: {str(e)}")
        return None

def laporan_stok(request):
    # Definisikan batas stok menipis (misalnya 5)
    BATAS_STOK_MENIPIS = 5  

    # Query untuk produk dengan stok menipis (lebih dari 0 tapi kurang dari atau sama dengan batas)
    stok_menipis = Produk.objects.filter(stok__gt=0, stok__lte=BATAS_STOK_MENIPIS).count()
    
    # Query untuk produk dengan stok habis (sama dengan 0)
    stok_habis = Produk.objects.filter(stok=0).count()

    context = {
        'stok_menipis': stok_menipis,
        'stok_habis': stok_habis,
    }
    
    return render(request, 'kasir/laporan_stok.html', context)

def export_stock_excel(request):
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Stok"

    # Header
    headers = ['No', 'Nama Barang', 'Kategori', 'Stok', 'Harga Beli', 'Harga Jual']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Data
    products = Produk.objects.all().select_related('kategori').order_by('stok')
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=product.nama_barang)
        ws.cell(row=row, column=3, value=product.kategori.nama if product.kategori else '-')
        ws.cell(row=row, column=4, value=product.stok)
        ws.cell(row=row, column=5, value=float(product.hp_beli))
        ws.cell(row=row, column=6, value=float(product.h_jual))

    # Styling
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Laporan_Stok.xlsx'
    wb.save(response)
    return response

def export_stock_pdf(request):
    # Get data
    products = Produk.objects.all().select_related('kategori').order_by('stok')
    
    # Get company info
    company_info = get_company_info()
    
    # Get logo dari StoreSettings
    store_settings = StoreSettings.objects.first()
    logo_base64 = get_logo_base64() if store_settings and store_settings.logo else None
    
    # Prepare context
    context = {
        'products': products,
        'today': datetime.now(),
        'request': request,
        'company_logo': logo_base64,
        'company_name': company_info['name'],
        'company_address': company_info['address'],
        'company_phone': company_info['phone'],
    }
    
    # Get template
    template = get_template('kasir/pdf_template.html')
    html = template.render(context)
    
    # Create PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f'Laporan_Stok_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse('Error Generating PDF', status=500)

def generate_cashflow_pdf(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        logger.debug(f"Generating cashflow PDF for period: {start_date} to {end_date}")
        
        if not start_date or not end_date:
            logger.error("Missing date parameters")
            return JsonResponse({
                'error': 'Missing date parameters'
            }, status=400)
        
        # Convert string dates to datetime
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Get cash flow data
        data = calculate_cashflow_data(start_date, end_date)
        logger.debug(f"Cashflow data calculated: {data}")
        
        # Get company info
        company_info = get_company_info()
        logger.debug(f"Company info retrieved: {company_info}")
        
        # Prepare context
        context = {
            'data': data,
            'start_date': start_date,
            'end_date': end_date,
            'current_datetime': datetime.now(),
            'company_name': company_info.get('name', 'Nama Toko'),
            'company_address': company_info.get('address', 'Alamat Toko'),
            'company_phone': company_info.get('phone', 'No. Telp'),
            'logo_base64': get_logo_base64(),
        }
        
        # Render template
        template = get_template('kasir/cashflow_pdf.html')
        html = template.render(context)
        
        # Create PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Laporan_Arus_Kas_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.pdf"'
        
        # Generate PDF
        pisa_status = pisa.CreatePDF(html, dest=response)
        
        if pisa_status.err:
            logger.error(f"Error generating PDF: {pisa_status.err}")
            return JsonResponse({
                'error': 'Failed to generate PDF'
            }, status=500)
        
        logger.debug("PDF generated successfully")
        return response
        
    except Exception as e:
        logger.exception("Error in generate_cashflow_pdf")
        return JsonResponse({
            'error': str(e)
        }, status=500)

def calculate_cashflow_data(start_date, end_date):
    """Calculate cash flow data for the given period"""
    try:
        # Penerimaan dari pelanggan (dari transaksi penjualan)
        cash_from_customers = Transaksi.objects.filter(
            tanggal__range=[start_date, end_date]
        ).aggregate(
            total=Coalesce(Sum('total'), Decimal('0'))
        )['total']
        
        # Pembayaran ke pemasok (dari pembelian stok)
        cash_to_suppliers = Produk.objects.filter(
            tgl_stok_masuk__range=[start_date, end_date]
        ).aggregate(
            total=Coalesce(Sum(F('stok') * F('hp_beli')), Decimal('0'))
        )['total']
        
        # Beban operasional (bisa disesuaikan dengan data real)
        operating_expenses = Decimal('0')
        
        # Hitung arus kas operasi
        net_operating_cash = cash_from_customers - cash_to_suppliers - operating_expenses
        
        # Aktivitas investasi (contoh, bisa disesuaikan)
        fixed_asset_purchase = Decimal('0')
        net_investing_cash = -fixed_asset_purchase
        
        # Aktivitas pendanaan (contoh, bisa disesuaikan)
        loan_received = Decimal('0')
        loan_payment = Decimal('0')
        net_financing_cash = loan_received - loan_payment
        
        # Perubahan kas bersih
        net_cash_change = net_operating_cash + net_investing_cash + net_financing_cash
        
        # Kas awal periode (bisa disesuaikan dengan data real)
        beginning_cash = Decimal('0')
        
        # Kas akhir periode
        ending_cash = beginning_cash + net_cash_change
        
        return {
            'cash_from_customers': cash_from_customers,
            'cash_to_suppliers': cash_to_suppliers,
            'operating_expenses': operating_expenses,
            'net_operating_cash': net_operating_cash,
            'fixed_asset_purchase': fixed_asset_purchase,
            'net_investing_cash': net_investing_cash,
            'loan_received': loan_received,
            'loan_payment': loan_payment,
            'net_financing_cash': net_financing_cash,
            'net_cash_change': net_cash_change,
            'beginning_cash': beginning_cash,
            'ending_cash': ending_cash
        }
        
    except Exception as e:
        print(f"Error in calculate_cashflow_data: {str(e)}")
        raise e

def get_company_info():
    """Get company information from settings or database"""
    # Ambil dari model StoreSettings jika ada
    try:
        store_settings = StoreSettings.objects.first()
        return {
            'name': store_settings.name,
            'address': store_settings.address,
            'phone': store_settings.phone
        }
    except:
        # Default values jika tidak ada di database
        return {
            'name': 'Nama Toko Anda',
            'address': 'Alamat Toko',
            'phone': 'No. Telepon'
        }

def get_logo_base64():
    try:
        # Coba ambil dari model StoreSettings
        store_settings = StoreSettings.objects.first()
        if store_settings and store_settings.logo:
            with open(store_settings.logo.path, 'rb') as image_file:
                return base64.b64encode(image_file.read()).decode()
                
        print("Logo tidak ditemukan di StoreSettings")
        return None
        
    except Exception as e:
        print(f"Error saat memuat logo: {str(e)}")
        return None

@require_http_methods(["POST"])
def save_module(request):
    try:
        module_id = request.POST.get('module_id')
        module_data = {
            'name': request.POST.get('module_name'),
            'path': request.POST.get('module_path'),
            'icon': request.POST.get('module_icon'),
            'order': request.POST.get('module_order'),
            'template': request.POST.get('module_template'),
            'active': request.POST.get('module_active') == 'on',
            'access_admin': request.POST.get('access_admin') == 'on',
            'access_kasir': request.POST.get('access_kasir') == 'on'
        }

        if module_id:
            Module.objects.filter(id=module_id).update(**module_data)
        else:
            Module.objects.create(**module_data)

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_http_methods(["GET"])
def get_module(request):
    try:
        module_id = request.GET.get('id')
        module = Module.objects.get(id=module_id)
        return JsonResponse({
            'id': module.id,
            'name': module.name,
            'path': module.path,
            'icon': module.icon,
            'order': module.order,
            'template': module.template,
            'active': module.active,
            'access_admin': module.access_admin,
            'access_kasir': module.access_kasir
        })
    except Module.DoesNotExist:
        return JsonResponse({'error': 'Module not found'}, status=404)

@require_http_methods(["POST"])
def toggle_module(request):
    try:
        module_id = request.POST.get('module_id')
        status = request.POST.get('status') == 'true'
        
        # Debug print
        print(f"Received: module_id={module_id}, status={status}")
        
        if not module_id:
            return JsonResponse({
                'status': 'error',
                'message': 'Module ID tidak ditemukan'
            }, status=400)

        # Jika module_id adalah 'ppob', handle secara khusus
        if module_id == 'ppob':
            # Simpan status ke dalam settings atau update modul PPOB yang sudah ada
            module, created = Module.objects.get_or_create(
                path='ppob',
                defaults={
                    'name': 'PPOB',
                    'icon': 'fa-mobile-alt',
                    'order': 1,
                    'template': 'blank',
                    'access_admin': True,
                    'access_kasir': True
                }
            )
            module.active = status
            module.save()
        else:
            # Handle modul lainnya
            module = Module.objects.get(id=module_id)
            module.active = status
            module.save()
        
        return JsonResponse({
            'status': 'success',
            'message': f'Status modul berhasil diubah menjadi {"aktif" if status else "nonaktif"}'
        })
        
    except Module.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Modul tidak ditemukan'
        }, status=404)
        
    except Exception as e:
        print(f"Error in toggle_module: {str(e)}")  # Debug print
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def ppob_view(request):
    # Get saldo from PPOBSaldo model
    saldo_obj = PPOBSaldo.objects.first()
    saldo = saldo_obj.balance if saldo_obj else 0
    
    context = {
        'title': 'PPOB',
        'saldo': saldo,
        'saldo_history': PPOBSaldoHistory.objects.all()[:10]
    }
    return render(request, 'kasir/modules/ppob.html', context)

@login_required
@require_POST
def ppob_add_saldo(request):
    try:
        nominal = Decimal(request.POST.get('nominal', '0'))
        keterangan = request.POST.get('keterangan', '')
        
        if nominal <= 0:
            return JsonResponse({
                'status': 'error',
                'message': 'Nominal harus lebih dari 0'
            })
            
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            if not saldo:
                saldo = PPOBSaldo.objects.create()
            
            balance_before = saldo.balance
            saldo.balance += nominal
            saldo.save()
            
            # Catat history
            PPOBSaldoHistory.objects.create(
                amount=nominal,
                balance_before=balance_before,
                balance_after=saldo.balance,
                type='IN',
                description=keterangan,
                created_by=request.user
            )
            
        return JsonResponse({
            'status': 'success',
            'message': 'Saldo berhasil ditambahkan',
            'new_balance': float(saldo.balance)
        })
        
    except ValueError:
        return JsonResponse({
            'status': 'error',
            'message': 'Nominal tidak valid'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
@require_POST
def ppob_beli_pulsa(request):
    try:
        phone = request.POST.get('phone')
        nominal = Decimal(request.POST.get('nominal', '0'))
        provider = request.POST.get('provider')
        
        # Ambil fee sesuai provider
        provider_fees = {
            'TELKOMSEL': {
                'base_fee': Decimal('2000'),
                'admin_fee': Decimal('2800'),
                'total_fee': Decimal('4800'),
                'mitra_fee': Decimal('450'),
                'agent_fee': Decimal('2250')
            },
            'INDOSAT': {
                'base_fee': Decimal('2000'),
                'admin_fee': Decimal('0'),
                'total_fee': Decimal('2000'),
                'mitra_fee': Decimal('475'),
                'agent_fee': Decimal('2375')
            },
            'XL': {
                'base_fee': Decimal('2000'),
                'admin_fee': Decimal('0'),
                'total_fee': Decimal('2000'),
                'mitra_fee': Decimal('425'),
                'agent_fee': Decimal('2125')
            },
            'TRI': {
                'base_fee': Decimal('2000'),
                'admin_fee': Decimal('0'),
                'total_fee': Decimal('2000'),
                'mitra_fee': Decimal('600'),
                'agent_fee': Decimal('3000')
            },
            'SMARTFREN': {
                'base_fee': Decimal('2000'),
                'admin_fee': Decimal('0'),
                'total_fee': Decimal('2000'),
                'mitra_fee': Decimal('600'),
                'agent_fee': Decimal('3000')
            }
        }
        
        fee = provider_fees.get(provider, {
            'base_fee': Decimal('0'),
            'admin_fee': Decimal('0'),
            'total_fee': Decimal('0'),
            'mitra_fee': Decimal('0'),
            'agent_fee': Decimal('0')
        })
        
        total_amount = nominal + fee['total_fee']
        
        # Proses transaksi
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            if not saldo or saldo.balance < total_amount:
                raise ValueError('Saldo tidak mencukupi')
            
            # Kurangi saldo
            saldo.balance -= total_amount
            saldo.save()
            
            # Catat history
            PPOBSaldoHistory.objects.create(
                amount=total_amount,
                balance_before=saldo.balance + total_amount,
                balance_after=saldo.balance,
                type='OUT',
                description=f'Pembelian pulsa {provider} {phone}',
                created_by=request.user
            )
            
        return JsonResponse({
            'status': 'success',
            'message': 'Pulsa berhasil dibeli',
            'new_balance': float(saldo.balance)
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
@require_POST
def ppob_beli_paket_data(request):
    try:
        phone = request.POST.get('phone')
        provider = request.POST.get('provider')
        paket = request.POST.get('paket')
        harga_str = request.POST.get('harga', '0')
        
        # Debug print
        print(f"Received data: phone={phone}, provider={provider}, paket={paket}, harga={harga_str}")
        
        try:
            # Konversi harga ke Decimal dengan handling yang lebih baik
            harga = decimal.Decimal(harga_str.replace(',', '').replace('.', ''))
        except decimal.InvalidOperation:
            raise ValueError(f'Harga tidak valid: {harga_str}')
            
        # Validasi input
        if not all([phone, provider, paket, harga]):
            raise ValueError('Data tidak lengkap')
            
        # Validasi saldo
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            
            print(f"Current balance: {saldo.balance if saldo else 'No saldo'}")
            
            if not saldo:
                raise ValueError('Saldo PPOB belum diinisialisasi')
                
            if saldo.balance < harga:
                raise ValueError(f'Saldo tidak mencukupi. Saldo: {saldo.balance}, Harga: {harga}')
            
            # Kurangi saldo
            saldo.balance -= harga
            saldo.save()
            
            # Catat history
            PPOBSaldoHistory.objects.create(
                amount=harga,
                balance_before=saldo.balance + harga,
                balance_after=saldo.balance,
                type='OUT',
                description=f'Pembelian paket data {provider} {phone} - {paket}',
                created_by=request.user
            )
            
        return JsonResponse({
            'status': 'success',
            'message': 'Paket data berhasil dibeli',
            'new_balance': float(saldo.balance)
        })
    except ValueError as e:
        print(f"ValueError in ppob_beli_paket_data: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
        
    except Exception as e:
        print(f"Unexpected error in ppob_beli_paket_data: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Terjadi kesalahan sistem'
        }, status=500)

@login_required
@require_POST
def ppob_beli_token(request):
    try:
        meter_number = request.POST.get('meter_number')
        nominal = Decimal(request.POST.get('nominal', '0'))
        
        # Fee untuk token listrik
        fee = {
            'admin_fee': Decimal('3000'),
            'total_fee': Decimal('3000'),
            'mitra_fee': Decimal('270'),
            'agent_fee': Decimal('1350')
        }
        
        total_amount = nominal + fee['total_fee']
        
        # Proses transaksi
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            if not saldo or saldo.balance < total_amount:
                raise ValueError('Saldo tidak mencukupi')
            
            # Kurangi saldo
            saldo.balance -= total_amount
            saldo.save()
            
            # Catat history
            PPOBSaldoHistory.objects.create(
                amount=total_amount,
                balance_before=saldo.balance + total_amount,
                balance_after=saldo.balance,
                type='OUT',
                description=f'Pembelian token listrik {meter_number}',
                created_by=request.user
            )
        return JsonResponse({
            'status': 'success',
            'message': 'Token listrik berhasil dibeli',
            'new_balance': float(saldo.balance)
        })
        
    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
        
    except Exception as e:
        print(f"Error in ppob_beli_token: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Terjadi kesalahan sistem'
        }, status=500)

@login_required
@require_POST
def ppob_bayar_pdam(request):
    try:
        customer_number = request.POST.get('customer_number')
        area = request.POST.get('area')
        nominal = decimal.Decimal(request.POST.get('nominal', '0'))
        admin_fee = decimal.Decimal(request.POST.get('admin_fee', '0'))
        total = decimal.Decimal(request.POST.get('total', '0'))
        
        # Validasi input
        if not all([customer_number, area, nominal, admin_fee, total]):
            raise ValueError('Data tidak lengkap')
            
        # Validasi saldo
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            
            if not saldo:
                raise ValueError('Saldo PPOB belum diinisialisasi')
                
            # Hanya mengecek nominal tagihan saja (tanpa biaya admin)
            if saldo.balance < nominal:
                raise ValueError(f'Saldo tidak mencukupi. Saldo: {saldo.balance}, Nominal: {nominal}')
            
            # Kurangi saldo hanya dengan nominal tagihan
            saldo.balance -= nominal
            saldo.save()
            
            # Catat history dengan detail biaya admin terpisah
            PPOBSaldoHistory.objects.create(
                amount=nominal,  # Hanya nominal tagihan
                balance_before=saldo.balance + nominal,
                balance_after=saldo.balance,
                type='OUT',
                description=f'Pembayaran PDAM {area} {customer_number} - Tagihan: Rp {nominal:,.0f} + Admin: Rp {admin_fee:,.0f}',
                created_by=request.user
            )
            
        return JsonResponse({
            'status': 'success',
            'message': 'Pembayaran PDAM berhasil',
            'new_balance': float(saldo.balance)
        })
        
    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
        
    except Exception as e:
        print(f"Error in ppob_bayar_pdam: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Terjadi kesalahan sistem'
        }, status=500)

@login_required
@require_POST
def ppob_tarik_tunai(request):
    try:
        customer_name = request.POST.get('customer_name')
        bank = request.POST.get('bank')
        nominal = decimal.Decimal(request.POST.get('nominal', '0'))
        
        # Validasi input
        if not all([customer_name, bank, nominal]):
            raise ValueError('Data tidak lengkap')
            
        # Proses penambahan saldo
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            
            if not saldo:
                raise ValueError('Saldo PPOB belum diinisialisasi')
            
            # Tambah saldo
            balance_before = saldo.balance
            saldo.balance += nominal
            saldo.save()
            
            # Catat history sebagai pemasukan
            PPOBSaldoHistory.objects.create(
                amount=nominal,
                balance_before=balance_before,
                balance_after=saldo.balance,
                type='IN',  # Ubah menjadi IN karena menambah saldo
                description=f'Tarik Tunai {bank} a.n {customer_name} - Rp {nominal:,.0f}',
                created_by=request.user
            )
            
        return JsonResponse({
            'status': 'success',
            'message': 'Tarik tunai berhasil',
            'new_balance': float(saldo.balance)
        })
        
    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
        
    except Exception as e:
        print(f"Error in ppob_tarik_tunai: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Terjadi kesalahan sistem'
        }, status=500)

@login_required
@require_POST
def ppob_transfer(request):
    try:
        customer_name = request.POST.get('customer_name')
        bank = request.POST.get('bank')
        nominal = decimal.Decimal(request.POST.get('nominal', '0'))
        
        # Fee untuk transfer antar bank
        fee = {
            'base_fee': Decimal('3500'),  # > Rp 1.000.000
            'admin_fee': Decimal('6500'),
            'total_fee': Decimal('10000'),
            'mitra_fee': Decimal('730'),
            'agent_fee': Decimal('3650')
        }
        
        # Jika nominal <= 1.000.000, gunakan fee yang berbeda
        if nominal <= 1000000:
            fee = {
                'base_fee': Decimal('2500'),
                'admin_fee': Decimal('6500'),
                'total_fee': Decimal('9000'),
                'mitra_fee': Decimal('630'),
                'agent_fee': Decimal('3150')
            }
        
        total_amount = nominal + fee['total_fee']
        
        # Validasi input dan proses transaksi
        # ... kode yang sudah ada ...

        # Validasi saldo
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            
            if not saldo:
                raise ValueError('Saldo PPOB belum diinisialisasi')
                
            if saldo.balance < nominal:
                raise ValueError(f'Saldo tidak mencukupi. Saldo: {saldo.balance}, Nominal: {nominal}')
            
            # Kurangi saldo
            saldo.balance -= nominal
            saldo.save()
            
            # Catat history
            PPOBSaldoHistory.objects.create(
                amount=nominal,
                balance_before=saldo.balance + nominal,
                balance_after=saldo.balance,
                type='OUT',
                description=f'Transfer {bank} a.n {customer_name} - Rp {nominal:,.0f}',
                created_by=request.user
            )
            
        return JsonResponse({
            'status': 'success',
            'message': 'Transfer berhasil',
            'new_balance': float(saldo.balance)
        })
        
    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
        
    except Exception as e:
        print(f"Error in ppob_transfer: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Terjadi kesalahan sistem'
        }, status=500)

@login_required
def ppob_chart_data(request):
    try:
        end_date = timezone.localtime(timezone.now()).date()
        start_date = end_date - timedelta(days=6)
        
        print(f"Querying data from {start_date} to {end_date}")
        
        # Query langsung dari database untuk debug
        with connection.cursor() as cursor:
            # Query untuk IN
            cursor.execute("""
                SELECT DATE(created_at) as date, SUM(amount) as total 
                FROM kasir_ppobsaldohistory 
                WHERE type = 'IN' 
                AND DATE(created_at) BETWEEN %s AND %s 
                GROUP BY DATE(created_at)
                ORDER BY date
            """, [start_date, end_date])
            in_data = dict(cursor.fetchall())
            
            # Query untuk OUT
            cursor.execute("""
                SELECT DATE(created_at) as date, SUM(amount) as total 
                FROM kasir_ppobsaldohistory 
                WHERE type = 'OUT' 
                AND DATE(created_at) BETWEEN %s AND %s 
                GROUP BY DATE(created_at)
                ORDER BY date
            """, [start_date, end_date])
            out_data = dict(cursor.fetchall())
        
        print("Raw IN data:", in_data)
        print("Raw OUT data:", out_data)
        
        # Generate data untuk setiap hari
        dates = []
        in_values = []
        out_values = []
        
        current = start_date
        while current <= end_date:
            dates.append(current.strftime('%d/%m'))
            in_values.append(float(in_data.get(current, 0)))
            out_values.append(float(out_data.get(current, 0)))
            current += timedelta(days=1)
        
        print("=== Final Data ===")
        print("Dates:", dates)
        print("IN values:", in_values)
        print("OUT values:", out_values)
        
        return JsonResponse({
            'labels': dates,
            'in_data': in_values,
            'out_data': out_values
        })
        
    except Exception as e:
        print(f"Error: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def ppob_today_summary(request):
    try:
        today = timezone.now().date()
        print(f"Getting summary for date: {today}")  # Debug print
        
        # Query untuk total masuk dan keluar hari ini
        with connection.cursor() as cursor:
            # Query untuk transaksi masuk (IN)
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) as total_in
                FROM kasir_ppobsaldohistory 
                WHERE DATE(created_at) = %s AND type = 'IN'
            """, [today])
            total_in = cursor.fetchone()[0]
            
            # Query untuk transaksi keluar (OUT)
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) as total_out
                FROM kasir_ppobsaldohistory 
                WHERE DATE(created_at) = %s AND type = 'OUT'
            """, [today])
            total_out = cursor.fetchone()[0]
            
            # Ambil saldo terkini
            cursor.execute("""
                SELECT balance_after 
                FROM kasir_ppobsaldohistory 
                ORDER BY created_at DESC 
                LIMIT 1
            """)
            current_balance = cursor.fetchone()
            
            print(f"Found: in={total_in}, out={total_out}, balance={current_balance}")  # Debug print
            
            return JsonResponse({
                'total_in': float(total_in),
                'total_out': float(total_out),
                'balance': float(current_balance[0] if current_balance else 0)
            })
            
    except Exception as e:
        print(f"Error in ppob_today_summary: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")
        return JsonResponse({
            'total_in': 0,
            'total_out': 0,
            'balance': 0
        })

@login_required
def ppob_export(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date or not end_date:
            raise ValueError("Tanggal harus diisi")

        # Debug print
        print(f"Exporting PPOB data from {start_date} to {end_date}")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan PPOB"
        
        # Headers
        headers = ['Tanggal', 'Waktu', 'Tipe', 'Keterangan', 'Masuk', 'Keluar', 'Saldo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="CCCCCC")
        
        # Ambil saldo awal (transaksi pertama/inisialisasi)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    amount as saldo_awal,
                    created_at,
                    description
                FROM kasir_ppobsaldohistory 
                WHERE type = 'IN'  -- Ambil transaksi masuk
                ORDER BY created_at ASC  -- Urutkan dari yang paling awal
                LIMIT 1  -- Ambil yang pertama
            """)
            
            saldo_awal_data = cursor.fetchone()
            if saldo_awal_data:
                saldo_awal = Decimal(str(saldo_awal_data[0]))
                saldo_awal_date = saldo_awal_data[1]
                saldo_awal_desc = saldo_awal_data[2]
                
                # Tulis saldo awal
                row = 2
                ws.cell(row=row, column=1, value=saldo_awal_date.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=2, value=saldo_awal_date.strftime('%H:%M:%S'))
                ws.cell(row=row, column=3, value='Info')
                ws.cell(row=row, column=4, value=f'Saldo Awal - {saldo_awal_desc}')
                ws.cell(row=row, column=5, value=float(saldo_awal))
                ws.cell(row=row, column=6, value=0)
                ws.cell(row=row, column=7, value=float(saldo_awal))
                
                # Format angka
                for col in [5, 6, 7]:
                    ws.cell(row=row, column=col).number_format = '#,##0'
                
                row += 1
            else:
                saldo_awal = Decimal('0')
                row = 2

            # Query transaksi sesuai periode, KECUALI transaksi saldo awal
            cursor.execute("""
                SELECT 
                    created_at,
                    type,
                    description,
                    amount,
                    balance_after
                FROM kasir_ppobsaldohistory 
                WHERE DATE(created_at) BETWEEN %s AND %s
                AND created_at > (SELECT created_at FROM kasir_ppobsaldohistory ORDER BY created_at ASC LIMIT 1)
                ORDER BY created_at
            """, [start_date, end_date])
            
            transactions = cursor.fetchall()

        total_in = Decimal('0')
        total_out = Decimal('0')
        
        # Proses transaksi selain saldo awal
        for trx in transactions:
            created_at = trx[0]
            trx_type = trx[1]
            description = trx[2]
            amount = Decimal(str(trx[3]))
            balance = Decimal(str(trx[4]))
            
            # Tanggal dan waktu
            ws.cell(row=row, column=1, value=created_at.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=created_at.strftime('%H:%M:%S'))
            ws.cell(row=row, column=3, value='Masuk' if trx_type == 'IN' else 'Keluar')
            ws.cell(row=row, column=4, value=description)
            
            if trx_type == 'IN':
                ws.cell(row=row, column=5, value=float(amount))
                ws.cell(row=row, column=6, value=0)
                total_in += amount
            else:
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=float(amount))
                total_out += amount
            
            ws.cell(row=row, column=7, value=float(balance))
            
            # Format angka
            for col in [5, 6, 7]:
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0'
            
            row += 1
        
        # Summary - sekarang tidak menghitung ulang saldo awal
        summary_row = row + 1
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=float(total_in)).font = Font(bold=True)
        ws.cell(row=summary_row, column=6, value=float(total_out)).font = Font(bold=True)
        ws.cell(row=summary_row, column=7, value=float(balance)).font = Font(bold=True)  # Gunakan balance terakhir
        
        # Format summary
        for col in [5, 6, 7]:
            ws.cell(row=summary_row, column=col).number_format = '#,##0'
        
        # Atur lebar kolom
        ws.column_dimensions['A'].width = 12  # Tanggal
        ws.column_dimensions['B'].width = 10  # Waktu
        ws.column_dimensions['C'].width = 8   # Tipe
        ws.column_dimensions['D'].width = 40  # Keterangan
        ws.column_dimensions['E'].width = 15  # Masuk
        ws.column_dimensions['F'].width = 15  # Keluar
        ws.column_dimensions['G'].width = 15  # Saldo
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Laporan_PPOB_{start_date}_{end_date}.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error in ppob_export: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")
        messages.error(request, f"Gagal mengexport data: {str(e)}")
        return redirect('kasir:ppob')

@login_required
@require_POST
def ppob_ewallet(request):
    try:
        phone = request.POST.get('phone')
        ewallet_type = request.POST.get('type')
        nominal = Decimal(request.POST.get('nominal', '0'))
        
        # Validasi input
        if not all([phone, ewallet_type, nominal]):
            raise ValueError('Data tidak lengkap')
            
        # Proses pengurangan saldo
        with transaction.atomic():
            saldo = PPOBSaldo.objects.select_for_update().first()
            
            if not saldo:
                raise ValueError('Saldo PPOB belum diinisialisasi')
            
            # Validasi saldo cukup
            if saldo.balance < nominal:
                raise ValueError('Saldo tidak mencukupi')
            
            # Kurangi saldo
            balance_before = saldo.balance
            saldo.balance -= nominal  # Kurangi saldo karena kita top up ke e-wallet customer
            saldo.save()
            
            # Catat history sebagai pengeluaran
            PPOBSaldoHistory.objects.create(
                amount=nominal,
                balance_before=balance_before,
                balance_after=saldo.balance,
                type='OUT',  # Pengeluaran karena kita top up ke e-wallet customer
                description=f'Top up {ewallet_type} {phone}',
                created_by=request.user
            )
            
        return JsonResponse({
            'status': 'success',
            'message': 'Top up e-wallet berhasil',
            'new_balance': float(saldo.balance)
        })
        
    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
        
    except Exception as e:
        print(f"Error in ppob_ewallet: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Terjadi kesalahan sistem'
        }, status=500)

@login_required
def module_view(request, module_path):
    """Generic view for handling dynamic modules"""
    # Get module or return 404
    module = get_object_or_404(Module, path=module_path, active=True)
    
    # Cek hak akses
    if request.user.is_superuser and not module.access_admin:
        raise Http404("Module not accessible")
    if not request.user.is_superuser and not module.access_kasir:
        raise Http404("Module not accessible")
    
    # Tentukan template yang akan digunakan
    template_name = f'kasir/modules/{module.path}.html'
    
    context = {
        'module': module,
        'title': module.name,
        'active_modules': Module.objects.filter(active=True).order_by('order')
    }
    
    return render(request, template_name, context)

@login_required
def ppob_get_data(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date or not end_date:
            return JsonResponse({
                'status': 'error',
                'message': 'Periode tanggal harus diisi'
            }, status=400)

        # Ambil data transaksi
        transactions = PPOBSaldoHistory.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).order_by('created_at')

        # Format data untuk chart
        chart_data = {
            'labels': [],
            'masuk': [],
            'keluar': []
        }

        # Hitung total per hari
        daily_data = {}
        for trx in transactions:
            date_str = trx.created_at.strftime('%d/%m/%Y')
            if date_str not in daily_data:
                daily_data[date_str] = {'masuk': 0, 'keluar': 0}
            
            if trx.type == 'IN':
                daily_data[date_str]['masuk'] += float(trx.amount)
            else:
                daily_data[date_str]['keluar'] += float(trx.amount)

        # Sort by date and populate chart data
        for date in sorted(daily_data.keys()):
            chart_data['labels'].append(date)
            chart_data['masuk'].append(daily_data[date]['masuk'])
            chart_data['keluar'].append(daily_data[date]['keluar'])

        return JsonResponse({
            'status': 'success',
            'data': chart_data
        })

    except Exception as e:
        print(f"Error in ppob_get_data: {str(e)}")  # Untuk debugging
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
@require_POST
def ppob_bayar_tagihan(request):
    try:
        jenis_tagihan = request.POST.get('jenis_tagihan')
        no_pelanggan = request.POST.get('no_pelanggan')
        nominal = decimal.Decimal(request.POST.get('nominal', '0'))
        
        # Fee berdasarkan jenis tagihan
        fee_config = {
            'TELKOM': {
                'base_fee': Decimal('1000'),
                'admin_fee': Decimal('2800'),
                'total_fee': Decimal('3800'),
                'mitra_fee': Decimal('350'),
                'agent_fee': Decimal('1750')
            },
            'INDIHOME': {
                'base_fee': Decimal('1000'),
                'admin_fee': Decimal('2800'),
                'total_fee': Decimal('3800'),
                'mitra_fee': Decimal('350'),
                'agent_fee': Decimal('1750')
            },
            'HP_PASCABAYAR': {
                'TELKOMSEL': {
                    'base_fee': Decimal('2000'),
                    'admin_fee': Decimal('2800'),
                    'total_fee': Decimal('4800'),
                    'mitra_fee': Decimal('450'),
                    'agent_fee': Decimal('2250')
                },
                'INDOSAT': {
                    'base_fee': Decimal('2000'),
                    'admin_fee': Decimal('0'),
                    'total_fee': Decimal('2000'),
                    'mitra_fee': Decimal('475'),
                    'agent_fee': Decimal('2375')
                },
                'XL': {
                    'base_fee': Decimal('2000'),
                    'admin_fee': Decimal('0'),
                    'total_fee': Decimal('2000'),
                    'mitra_fee': Decimal('425'),
                    'agent_fee': Decimal('2125')
                },
                'TRI': {
                    'base_fee': Decimal('2000'),
                    'admin_fee': Decimal('0'),
                    'total_fee': Decimal('2000'),
                    'mitra_fee': Decimal('600'),
                    'agent_fee': Decimal('3000')
                },
                'SMARTFREN': {
                    'base_fee': Decimal('2000'),
                    'admin_fee': Decimal('0'),
                    'total_fee': Decimal('2000'),
                    'mitra_fee': Decimal('600'),
                    'agent_fee': Decimal('3000')
                }
            }
        }
        
        # Ambil fee sesuai jenis tagihan
        fee = fee_config.get(jenis_tagihan, {
            'base_fee': Decimal('0'),
            'admin_fee': Decimal('0'),
            'total_fee': Decimal('0'),
            'mitra_fee': Decimal('0'),
            'agent_fee': Decimal('0')
        })
        
        total_amount = nominal + fee['total_fee']
        
        return JsonResponse({
            'status': 'success',
            'message': 'Tagihan berhasil dibayar',
            'total_amount': float(total_amount)
        })
        
    except Exception as e:
        print(f"Error in ppob_bayar_tagihan: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
@require_POST
def backup_database(request):
    try:
        # Dapatkan konfigurasi database dari settings
        db_settings = settings.DATABASES['default']
        
        # Buat nama file dengan timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_database_{timestamp}.sql'
        backup_file = os.path.join(settings.MEDIA_ROOT, 'backups', filename)
        
        # Pastikan direktori backup ada
        os.makedirs(os.path.dirname(backup_file), exist_ok=True)
        
        # Buat command untuk mysqldump
        cmd = [
            'mysqldump',
            '--user=' + db_settings['USER'],
            '--password=' + db_settings['PASSWORD'],
            '--host=' + db_settings['HOST'],
            '--port=' + str(db_settings['PORT']),
            db_settings['NAME']
        ]
        
        # Jalankan mysqldump dan simpan ke file
        with open(backup_file, 'w') as f:
            subprocess.run(cmd, stdout=f, check=True)
        
        # Baca file backup
        with open(backup_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/sql')
            response['Content-Disposition'] = f'attachment; filename={filename}'
        
        # Hapus file temporary
        os.remove(backup_file)
        
        return response
        
    except Exception as e:
        print(f"Backup error: {str(e)}")  # Untuk debugging
        return JsonResponse({
            'status': 'error',
            'message': f'Gagal membuat backup: {str(e)}'
        }, status=500)

@login_required
@require_POST
def restore_database(request):
    try:
        if 'restoreFile' not in request.FILES:
            return JsonResponse({
                'status': 'error',
                'message': 'File tidak ditemukan'
            }, status=400)
            
        sql_file = request.FILES['restoreFile']
        
        # Validasi ekstensi file
        if not sql_file.name.endswith('.sql'):
            return JsonResponse({
                'status': 'error',
                'message': 'File harus berformat .sql'
            }, status=400)
        
        # Simpan file temporary
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_file = os.path.join(settings.MEDIA_ROOT, 'temp', f'restore_{timestamp}.sql')
        os.makedirs(os.path.dirname(temp_file), exist_ok=True)
        
        with open(temp_file, 'wb+') as f:
            for chunk in sql_file.chunks():
                f.write(chunk)
        
        # Dapatkan konfigurasi database
        db_settings = settings.DATABASES['default']
        
        # Buat command untuk mysql
        cmd = [
            'mysql',
            '--user=' + db_settings['USER'],
            '--password=' + db_settings['PASSWORD'],
            '--host=' + db_settings['HOST'],
            '--port=' + str(db_settings['PORT']),
            db_settings['NAME'],
            '-e',
            f'source {temp_file}'
        ]
        
        # Jalankan restore
        subprocess.run(cmd, check=True)
        
        # Hapus file temporary
        os.remove(temp_file)
        
        return JsonResponse({
            'status': 'success',
            'message': 'Database berhasil direstore'
        })
        
    except subprocess.CalledProcessError as e:
        print(f"Restore error: {str(e)}")  # Untuk debugging
        return JsonResponse({
            'status': 'error',
            'message': f'Gagal melakukan restore: {str(e)}'
        }, status=500)
        
    except Exception as e:
        print(f"Restore error: {str(e)}")  # Untuk debugging
        return JsonResponse({
            'status': 'error',
            'message': f'Gagal melakukan restore: {str(e)}'
        }, status=500)

@require_POST
def delete_transaction(request, transaction_id):
    try:
        with transaction.atomic():
            # Ambil transaksi
            transaksi = get_object_or_404(Transaksi, id=transaction_id)
            
            # Kembalikan stok produk
            for detail in transaksi.transaksidetail_set.all():
                produk = detail.produk
                produk.stok += detail.qty
                produk.save()
            
            # Hapus transaksi (akan menghapus detail transaksi juga karena CASCADE)
            transaksi.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Transaksi #{transaction_id} berhasil dihapus'
            })
            
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

def fix_transaction_details(request):
    try:
        # Ambil semua transaksi detail yang produk_id-nya NULL
        null_details = TransaksiDetail.objects.filter(produk_id__isnull=True)
        updated_count = 0
        
        for detail in null_details:
            # Cari produk berdasarkan harga jual
            try:
                produk = Produk.objects.get(h_jual=detail.harga)
                detail.produk_id = produk.id
                detail.save()
                updated_count += 1
            except Produk.DoesNotExist:
                continue
            except Produk.MultipleObjectsReturned:
                # Jika ada beberapa produk dengan harga yang sama, ambil yang pertama
                produk = Produk.objects.filter(h_jual=detail.harga).first()
                detail.produk_id = produk.id
                detail.save()
                updated_count += 1
        
        return JsonResponse({
            'status': 'success',
            'message': f'Berhasil memperbaiki {updated_count} transaksi detail'
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

